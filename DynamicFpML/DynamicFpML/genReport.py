from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, Fill, NamedStyle, is_date_format

font = Font(name='Calibri', size=11, bold=True, italic=False, vertAlign=None, underline='none', strike=False,
            color='FF000000')
fill = PatternFill(fill_type="solid", start_color='FF008000', end_color='FF008000')
from datetime import datetime
import getpass, os, re

# for header
header = NamedStyle(name="header")
header.font = Font(name='Calibri', size=11, bold=True, italic=False, vertAlign=None, underline='none', strike=False,
                   color='FFFFFF')
header.fill = PatternFill(fill_type="solid", start_color='00ab4e', end_color='00ab4e')
header.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                       bottom=Side(style='medium'))
#  for even row
even_row = NamedStyle(name="odd_row")
even_row.font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None,
                    underline='none', strike=False,
                    color='000000')
even_row.fill = PatternFill(fill_type="solid", start_color='d1d1d1', end_color='d1d1d1')
even_row.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
# for odd row
odd_row = NamedStyle(name="even_row")
odd_row.font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None,
                    underline='none', strike=False,
                    color='000000')
odd_row.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
#  for pass_style
pass_style = NamedStyle(name="pass_style")
pass_style.font = Font(name='Calibri', size=11, bold=True, italic=False, vertAlign=None,
                       underline='none', strike=False,
                       color='000000')
pass_style.fill = PatternFill(fill_type="solid", start_color='99CC99', end_color='99CC99')
pass_style.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

# for fail_style
fail_style = NamedStyle(name="fail_style")
fail_style.font = Font(name='Calibri', size=11, bold=True, italic=False, vertAlign=None,
                       underline='none', strike=False,
                       color='000000')
fail_style.fill = PatternFill(fill_type="solid", start_color='ffc6c6', end_color='ffc6c6')
fail_style.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

filepath = ''
values_list = list()
values_list = [['Trade ID', 'Pass', 'Fail', 'Total Comparisons']]

def readfile():
    file = open(r"C:\Users\gaurav.saini\Desktop\DynamicFpML\DynamicFpML\DATA\CompareXML.log", "r")
    return file


def createWorkbook(logfile):
    global even_row, pass_style, fail_style
    currDateTime = datetime.now()
    filename = "CompareFpML_Result_{0}.xlsx".format(currDateTime.strftime('%d%b%Y-%H%M%S'))
    filename2 = "CompareFpML_Result.xlsx"
    wb = Workbook()
    ws1 = wb.active

    TicketName = "TradeSTP CitiPB bulk replay"
    ws1.title = "ResultTable"
    ws1['A1'] = "Project"
    ws1.cell(row=1, column=1).style = header
    ws1['B1'] = TicketName
    ws1.cell(row=1, column=2).style = even_row


    ws1['A2'] = "Tester"
    ws1.cell(row=2, column=1).style = header
    ws1['B2'] = os.environ['USERPROFILE'][9:]
    ws1.cell(row=2, column=2).style = even_row

    ws1['A3'] = "RunDate"
    ws1.cell(row=3, column=1).style = header
    ws1['A4'] = "RunTime"
    ws1.cell(row=4, column=1).style = header

    val = "{}".format(currDateTime.strftime("%d %B %Y; %A"))
    ws1['B3'] = val
    ws1.cell(row=3, column=2).style = even_row
    val = "{}".format(currDateTime.strftime("%I:%M %p IST"))
    ws1['B4'] = val
    ws1.cell(row=4, column=2).style = even_row

    ws1['A5'] = "Environment"
    ws1.cell(row=5, column=1).style = header
    ws1['B5'] = "QA"
    ws1.cell(row=5, column=2).style = even_row

    ws1['A6'] = "Test Result"
    ws1.cell(row=6, column=1).style = header

    ws1['A8'] = "TradeID"
    ws1.cell(row=8, column=1).style = header
    ws1['B8'] = "xPath"
    ws1.cell(row=8, column=2).style = header
    ws1['C8'] = "Expected"
    ws1.cell(row=8, column=3).style = header
    ws1['D8'] = "Actual"
    ws1.cell(row=8, column=4).style = header
    """
   # working without profile info
    ws1['A1'] = "TradeID"
    ws1.cell(row=1, column=1).style = header
    ws1['B1'] = "xPath"
    ws1.cell(row=1, column=2).style = header
    ws1['C1'] = "Expected"
    ws1.cell(row=1, column=3).style = header
    ws1['D1'] = "Actual"
    ws1.cell(row=1, column=4).style = header
    """

    ws2 = wb.create_sheet(title="ResultGraph")

    """
    ws2['A1'] = "Test Summary"
    ws2['A2'] = "Trade ID"
    ws2.cell(row=2, column=1).style = header
    ws2['B2'] = "Total"
    ws2.cell(row=2, column=2).style = header
    ws2['C2'] = "Pass"
    ws2.cell(row=2, column=3).style = header
    ws2['D2'] = "Fail"
    ws2.cell(row=2, column=4).style = header """

    ws1, result = addTables(ws=ws1, file=logfile)
    ws2 = addGraphTable(ws=ws2)

    ws1['B6'] = result
    if result == "Pass":
        ws1.cell(row=6, column=2).style = pass_style
    else:
        ws1.cell(row=6, column=2).style = fail_style


    global  values_list
    # print values_list

    # saveDIR = os.path.join(os.path.join(os.environ['USERPROFILE'], 'Desktop/CompareResult2/'))
    saveDIR = r'C:\Users\gaurav.saini\Desktop\DynamicFpML\DynamicFpML\DATA'
    if not os.path.exists(saveDIR):
        os.makedirs(saveDIR)
    global filepath
    filepath = saveDIR + '\\' + filename
    wb.save(filepath)
    sub_date = format(currDateTime.strftime('%d%b%Y'))
    return "saveDIR + filename2", filename2, sub_date


def addTables(ws, file):
    global values_list
    temp_list = list()
    with open(file, "r") as filelog:
        # filelog = readfile()
        lines = re.split(r'\n{1,}', filelog.read())
        lines = filter(None, lines)
        styleTem = even_row
        rowCount = 9
        colCount = 1
        tradeID = 1
        total_compared = 0
        total_fail = 0
        Trade_Count = 0
        result = ""
        for line in lines:
            if "QATestTradeID" in line:
                tradeID = re.search('QATestTradeID: (.*)', line)
                total_fail = 0
                Trade_Count += 1
                if Trade_Count % 2 == 0:
                    styleTem = even_row
                else:
                    styleTem = odd_row
            elif "Total_Compared: " in line:
                total_compared = re.search('Total_Compared: (.*)', line)

                temp_list.append(tradeID.group(1))
                temp_list.append((int(total_compared.group(1))) - total_fail)
                temp_list.append(total_fail)
                temp_list.append(int(total_compared.group(1)))
                values_list.append(temp_list)
                temp_list = []
                # values_list.extend([tradeID, total_compared, total_compared - total_fail, total_fail])
                styleTem = odd_row
                result = "Fail"

            else:
                total_fail += 1
                colCount = 1
                # """if rowCount % 2 == 0:
                #     styleTem = even_row
                # else:
                #     styleTem = odd_row"""

                ws.cell(row=rowCount, column=colCount).value = tradeID.group(1)
                ws.cell(row=rowCount, column=colCount).style = styleTem

                xpath = re.search('differs at (.*). Expected: ', line)

                ws.cell(row=rowCount, column=colCount+1).value = xpath.group(1)
                ws.cell(row=rowCount, column=colCount+1).style = styleTem

                expected = re.search('Expected: (.*) Actual: ', line)
                ws.cell(row=rowCount, column=colCount+2).value = expected.group(1)
                ws.cell(row=rowCount, column=colCount+2).style = styleTem

                actual = re.search('Actual: (.*)', line)
                ws.cell(row=rowCount, column=colCount+3).value = actual.group(1)
                ws.cell(row=rowCount, column=colCount+3).style = styleTem

                rowCount += 1
            filelog.close()

        return ws, result

def addGraphTable(ws):
    ws['A1'] = "Test Summary"
    counter = 2
    for i in values_list:
        if counter == 2:
            ws['A{0}'.format(counter)] = i[0]
            ws.cell(row=2, column=1).style = header
            ws['B{0}'.format(counter)] = i[1]
            ws.cell(row=2, column=2).style = header
            ws['C{0}'.format(counter)] = i[2]
            ws.cell(row=2, column=3).style = header
            ws['D{0}'.format(counter)] = i[3]
            ws.cell(row=2, column=4).style = header
        else:
            if counter % 2 == 0:
                styleTem = even_row
            else:
                styleTem = odd_row

            ws['A{0}'.format(counter)] = i[0]
            ws.cell(row=counter, column=1).style = styleTem
            ws['B{0}'.format(counter)] = i[1]
            ws.cell(row=counter, column=2).style = styleTem
            ws['C{0}'.format(counter)] = i[2]
            ws.cell(row=counter, column=3).style = styleTem
            ws['D{0}'.format(counter)] = i[3]
            ws.cell(row=counter, column=4).style = styleTem

        counter += 1

    row = 3
    chart = 1

    from openpyxl.chart import (
        PieChart,
        ProjectedPieChart,
        Reference
    )
    from openpyxl.chart.series import DataPoint

    for data in range(len(values_list) - 1):
        pie = PieChart()

        labels = Reference(ws, min_col=2, max_col=3, min_row=2, max_row=2)
        data = Reference(ws, min_col=2, max_col=3, min_row=row, max_row=row)

        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = values_list[row - 2][0]
        ws.add_chart(pie, "F{}".format(chart))
        row += 1
        chart += 14

    """for x in values_list:

        rr = i+1
        labels_1 = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=1)
        # data_p = Reference(ws, min_col=2, max_col=4, min_row=2, max_row=2)
        data_p = Reference(ws, min_col=2, max_col=3, min_row=rr, max_row=rr)

        pie.add_data(data_p, titles_from_data=True)
        pie.set_categories(labels_1)
        pie.title = "{}".format(x[0])
        ws.add_chart(pie, "F{}".format(i*19))
        i += 1"""
    return ws


if __name__ == "__main__":
    file = r"C:\Users\gaurav.saini\Desktop\DynamicFpML\DynamicFpML\DATA\CompareXML.log"
    path, filename, subdate = createWorkbook(file)

    # line = r'Difference 3: Element Text differs at //FpML/party/partyName. Expected: PROLOGUE Capital Actual: PROLOGUE Capital gain'
    # xpath = re.search('differs at (.*). Expected: ', line)
    # print xpath.group(1)
