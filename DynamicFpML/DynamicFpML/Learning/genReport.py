from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, Fill, NamedStyle
font = Font(name='Calibri', size=11, bold=True, italic=False, vertAlign=None, underline='none', strike=False,
            color='FF000000')
fill = PatternFill(fill_type="solid", start_color='FF008000', end_color='FF008000')
from datetime import datetime
import getpass, os, re

template_style_header = NamedStyle(name="template_style")
template_style_header.font = Font(name='Calibri', size=11, bold=True, italic=False, vertAlign=None, underline='none', strike=False,
            color='FFFFFFFF')
template_style_header.fill = PatternFill(fill_type="solid", start_color='FF008000', end_color='FF008000')
# template_style_header.fill = Border(left=Side(border_style="medium"),
#                 right=Side(border_style="medium", color='FF000000'),
#                 top=Side(border_style="medium", color='FF000000'),
#                 bottom=Side(border_style="medium", color='FF000000'), diagonal=Side(border_style= ))

def readfile():
    file = open(r"D:\MarkitServ\BAU\Python\xml_compare\xml-test\Learning\CompareXML.log", "r")
    return file

def createWorkbook(logfile):
    currDateTime = datetime.now()
    filename = "CompareXML_Result_{0}.xlsx".format(currDateTime.strftime('%d%b%Y-%H%M%S'))
    filename2 = "CompareXML_Result.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "ResultTable"
    ws1['A1'] = "TradeID"
    ws1.cell(row=1, column=1).style = template_style_header
    ws1['B1'] = "xPath"
    ws1.cell(row=1, column=2).style = template_style_header
    ws1['C1'] = "Expected"
    ws1.cell(row=1, column=3).style = template_style_header
    ws1['D1'] = "Actual"
    ws1.cell(row=1, column=4).style = template_style_header
    ws2 = wb.create_sheet(title="ResultGraph")
    ws1 = addTables(ws=ws1, file=logfile)
    wb.save(os.path.join(os.environ['USERPROFILE'], 'Desktop/CompareResult/', '{0}').format(filename2))


def addTables(ws, file):
    with open(file, "r") as filelog:
    #filelog = readfile()
        rowCount = 2
        colCount = 1
        tradeID = 1
        for line in filelog:
            if "QATestTradeID" in line:
                tradeID = re.search('QATestTradeID: (.*)', line)
            else:
                colCount = 1
                ws.cell(row=rowCount, column=colCount).value = tradeID.group(1)
                xpath = re.search('differs at (.*). Expected: ', line)
                ws.cell(row=rowCount, column=colCount + 1).value = xpath.group(1)
                expected = re.search('Expected: (.*) Actual: ', line)
                ws.cell(row=rowCount, column=colCount + 2).value = expected.group(1)
                actual = re.search('Actual: (.*)', line)
                ws.cell(row=rowCount, column=colCount + 3).value = actual.group(1)
                rowCount += 1
        return ws


if __name__ == "__main__":
    file = r"D:\MarkitServ\BAU\Python\xml_compare\xml-test\Learning\CompareXML.log"
    createWorkbook(file)
